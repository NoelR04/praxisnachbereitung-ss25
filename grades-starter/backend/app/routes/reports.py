from fastapi import APIRouter, Response
import csv
import io
from openpyxl import Workbook

from ..services.grade_report_service import (
  export_cell_value,
  get_grade_export_fieldnames,
  get_grade_export_rows,
  get_grade_report_rows,
)

router = APIRouter(tags=["reports"])


@router.get("/reports/grade-overview")
def grade_overview():
  return list(get_grade_report_rows())


@router.get("/reports/grade-overview.csv")
def grade_overview_csv():
  rows = list(get_grade_report_rows())
  if not rows:
    rows = [{
      "matrikel": "",
      "student_name": "",
      "programme": "",
      "semester": "",
      "module_name": "",
      "grade_value": "",
      "attempt": "",
      "graded_at": "",
    }]

  fieldnames = list(rows[0].keys())
  buf = io.StringIO()
  writer = csv.DictWriter(buf, fieldnames=fieldnames, delimiter=";", lineterminator="\n")
  writer.writeheader()
  for row in rows:
    writer.writerow(row)

  data = buf.getvalue().encode("utf-8-sig")
  headers = {"Content-Disposition": 'attachment; filename="grade-overview.csv"'}
  return Response(content=data, media_type="text/csv; charset=utf-8", headers=headers)


@router.get("/grades.csv")
def grades_csv():
  rows = get_grade_export_rows()
  buf = io.StringIO()
  fieldnames = get_grade_export_fieldnames()
  writer = csv.DictWriter(buf, fieldnames=fieldnames, delimiter=";", lineterminator="\n")
  writer.writeheader()
  for row in rows:
    writer.writerow(row)

  data = buf.getvalue().encode("utf-8-sig")
  headers = {"Content-Disposition": 'attachment; filename="grades.csv"'}
  return Response(content=data, media_type="text/csv; charset=utf-8", headers=headers)


@router.get("/grades.xlsx")
def grades_xlsx():
  rows = get_grade_export_rows()
  fieldnames = get_grade_export_fieldnames()
  grade_value_col = fieldnames.index("grade_value") + 1
  graded_at_col = fieldnames.index("graded_at") + 1

  workbook = Workbook()
  sheet = workbook.active
  sheet.title = "grades"
  sheet.append(fieldnames)

  for row_idx, row in enumerate(rows, start=2):
    sheet.append([export_cell_value(field, row[field]) for field in fieldnames])
    sheet.cell(row=row_idx, column=grade_value_col).number_format = "0.0#"
    sheet.cell(row=row_idx, column=graded_at_col).number_format = "DD.MM.YYYY HH:MM:SS"

  buf = io.BytesIO()
  workbook.save(buf)
  headers = {"Content-Disposition": 'attachment; filename="grades.xlsx"'}
  return Response(
    content=buf.getvalue(),
    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    headers=headers,
  )
